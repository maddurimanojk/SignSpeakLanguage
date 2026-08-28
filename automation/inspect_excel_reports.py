import os
import openpyxl

excel_dir = "automation/reports/Excel"
files = os.listdir(excel_dir)

print(f"=== INSPECTING {len(files)} GENERATED EXCEL WORKBOOKS ===")

for filename in sorted(files):
    if not filename.endswith(".xlsx"):
        continue
    filepath = os.path.join(excel_dir, filename)
    wb = openpyxl.load_workbook(filepath)
    print(f"\nWorkbook: {filename}")
    print(f"Sheet names: {wb.sheetnames}")
    
    ws = wb.active
    print(f"Active Sheet: '{ws.title}' (Rows: {ws.max_row}, Cols: {ws.max_column})")
    
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Headers: {headers}")
    
    # Print sample row 2 and row 3
    for r in range(2, min(4, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {r}: {row_vals}")
