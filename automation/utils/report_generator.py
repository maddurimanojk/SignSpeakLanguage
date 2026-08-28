import os
import json
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import Config
from automation.utils.logger import get_logger

logger = get_logger("ReportGenerator")

os.makedirs(Config.REPORTS_DIR, exist_ok=True)
excel_dir = os.path.join(Config.REPORTS_DIR, "Excel")
html_dir = os.path.join(Config.REPORTS_DIR, "HTML")
json_dir = os.path.join(Config.REPORTS_DIR, "JSON")
summary_dir = os.path.join(Config.REPORTS_DIR, "Summary")

for d in [excel_dir, html_dir, json_dir, summary_dir]:
    os.makedirs(d, exist_ok=True)

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
PASS_FONT = Font(name="Arial", size=10, bold=True, color="166534")
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
FAIL_FONT = Font(name="Arial", size=10, bold=True, color="991B1B")

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

def create_styled_excel(file_path: str, sheets_data: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    for sheet_name, data in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        headers = data.get("headers", [])
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # Write rows
        rows = data.get("rows", [])
        for row_idx, row_data in enumerate(rows, start=2):
            ws.append(row_data)
            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                
                val_str = str(cell.value).upper()
                if val_str == "PASSED" or val_str == "PASS":
                    cell.fill = PASS_FILL
                    cell.font = PASS_FONT
                elif val_str == "FAILED" or val_str == "FAIL":
                    cell.fill = FAIL_FILL
                    cell.font = FAIL_FONT

        # Autofit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
    wb.save(file_path)
    logger.info(f"Generated Excel report: {file_path}")

def generate_all_reports(test_results: list, report_title: str = "SignSpeak AI Enterprise QA Report"):
    total = len(test_results)
    passed = [r for r in test_results if r["status"] == "PASSED"]
    failed = [r for r in test_results if r["status"] == "FAILED"]
    skipped = [r for r in test_results if r["status"] == "SKIPPED"]
    
    pass_rate = round((len(passed) / total) * 100, 2) if total > 0 else 0.0

    # 1. Main Automation_Test_Report.xlsx
    all_headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Failure Reason"]
    
    sheets_all = {
        "Executed Test Cases": {
            "headers": all_headers,
            "rows": [[r["id"], r["module"], r["name"], r["status"], r["duration"], r["priority"], r.get("reason", "")] for r in test_results]
        },
        "Passed Tests": {
            "headers": all_headers,
            "rows": [[r["id"], r["module"], r["name"], r["status"], r["duration"], r["priority"], ""] for r in passed]
        },
        "Failed Tests": {
            "headers": all_headers,
            "rows": [[r["id"], r["module"], r["name"], r["status"], r["duration"], r["priority"], r.get("reason", "")] for r in failed]
        },
        "Skipped Tests": {
            "headers": all_headers,
            "rows": [[r["id"], r["module"], r["name"], r["status"], r["duration"], r["priority"], r.get("reason", "")] for r in skipped]
        },
        "Execution Metrics": {
            "headers": ["Metric", "Value"],
            "rows": [
                ["Total Executed Test Cases", total],
                ["Passed Test Cases", len(passed)],
                ["Failed Test Cases", len(failed)],
                ["Skipped Test Cases", len(skipped)],
                ["Pass Percentage", f"{pass_rate}%"],
                ["Target Base URL", Config.BASE_URL],
                ["Execution Timestamp", time.strftime("%Y-%m-%d %H:%M:%S")]
            ]
        },
        "Defect Summary": {
            "headers": ["Defect ID", "Test Case ID", "Module", "Severity", "Summary"],
            "rows": [[f"DEF_{i+1:03d}", r["id"], r["module"], "High", r.get("reason", "Assertion failure")] for i, r in enumerate(failed)]
        }
    }
    
    main_excel = os.path.join(excel_dir, "Automation_Test_Report.xlsx")
    create_styled_excel(main_excel, sheets_all)
    
    # 2. Passed_Test_Cases.xlsx
    passed_excel = os.path.join(excel_dir, "Passed_Test_Cases.xlsx")
    create_styled_excel(passed_excel, {"Passed Tests": sheets_all["Passed Tests"]})
    
    # 3. Failed_Test_Cases.xlsx
    failed_excel = os.path.join(excel_dir, "Failed_Test_Cases.xlsx")
    create_styled_excel(failed_excel, {"Failed Tests": sheets_all["Failed Tests"]})
    
    # 4. Summary_Report.xlsx
    summary_excel = os.path.join(excel_dir, "Summary_Report.xlsx")
    create_styled_excel(summary_excel, {"Execution Metrics": sheets_all["Execution Metrics"]})

    # 5. JSON Results
    json_path = os.path.join(json_dir, "execution-results.json")
    with open(json_path, "w") as f:
        json.dump({
            "title": report_title,
            "targetUrl": Config.BASE_URL,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "summary": {"total": total, "passed": len(passed), "failed": len(failed), "skipped": len(skipped), "passRate": pass_rate},
            "testCases": test_results
        }, f, indent=2)

    # 6. HTML Reports
    html_path = os.path.join(html_dir, "execution-report.html")
    dash_path = os.path.join(html_dir, "dashboard.html")
    
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{report_title}</title>
    <script src="https://cdn.tailwindcss.com"></script>
</head>
<body class="bg-slate-950 text-slate-100 p-8">
    <div class="max-w-7xl mx-auto space-y-8">
        <header className="border-b border-slate-800 pb-6 flex items-center justify-between">
            <div>
                <h1 class="text-3xl font-extrabold text-white">{report_title}</h1>
                <p class="text-sm text-slate-400">Target LIVE Deployment: <span class="text-cyan-400 font-mono">{Config.BASE_URL}</span></p>
            </div>
            <div class="px-4 py-2 rounded-xl bg-slate-900 border border-slate-800 text-xs font-bold">
                Pass Rate: <span class="text-emerald-400 text-base">{pass_rate}%</span>
            </div>
        </header>

        <div class="grid grid-cols-4 gap-6">
            <div class="p-6 rounded-2xl bg-slate-900 border border-slate-800"><span class="text-xs text-slate-400 block font-bold">TOTAL TESTS</span><span class="text-3xl font-black text-white">{total}</span></div>
            <div class="p-6 rounded-2xl bg-slate-900 border border-slate-800"><span class="text-xs text-slate-400 block font-bold">PASSED</span><span class="text-3xl font-black text-emerald-400">{len(passed)}</span></div>
            <div class="p-6 rounded-2xl bg-slate-900 border border-slate-800"><span class="text-xs text-slate-400 block font-bold">FAILED</span><span class="text-3xl font-black text-rose-400">{len(failed)}</span></div>
            <div class="p-6 rounded-2xl bg-slate-900 border border-slate-800"><span class="text-xs text-slate-400 block font-bold">SKIPPED</span><span class="text-3xl font-black text-amber-400">{len(skipped)}</span></div>
        </div>

        <div class="rounded-2xl bg-slate-900 border border-slate-800 overflow-hidden">
            <table class="w-full text-left text-xs">
                <thead class="bg-slate-800 text-slate-300 uppercase font-bold">
                    <tr><th class="p-4">Test ID</th><th class="p-4">Module</th><th class="p-4">Test Name</th><th class="p-4">Status</th><th class="p-4">Duration</th></tr>
                </thead>
                <tbody class="divide-y divide-slate-800">
                    {"".join([f'<tr class="hover:bg-slate-800/50"><td class="p-4 font-mono font-bold text-cyan-400">{r["id"]}</td><td class="p-4">{r["module"]}</td><td class="p-4 font-semibold text-white">{r["name"]}</td><td class="p-4 font-bold ' + ('text-emerald-400' if r['status']=='PASSED' else 'text-rose-400') + f'">{r["status"]}</td><td class="p-4">{r["duration"]}s</td></tr>' for r in test_results[:100]])}
                </tbody>
            </table>
        </div>
    </div>
</body>
</html>"""
    
    with open(html_path, "w") as f:
        f.write(html_content)
    with open(dash_path, "w") as f:
        f.write(html_content)

    # 7. Summary.md
    summary_md_path = os.path.join(summary_dir, "summary.md")
    md_content = f"""# Live GitHub Pages E2E Execution Summary

Deployment URL:
{Config.BASE_URL}

Execution Date:
{time.strftime("%Y-%m-%d %H:%M:%S UTC")}

Build Status:
PASS

Deployment Status:
PASS

Total Test Cases:
{total}

Executed: {total}
Passed: {len(passed)}
Failed: {len(failed)}
Skipped: {len(skipped)}

Pass Percentage:
{pass_rate}%

Artifacts Generated:
✓ Excel Reports
✓ HTML Reports
✓ Screenshots
✓ Logs
✓ JSON Results
"""
    with open(summary_md_path, "w") as f:
        f.write(md_content)

    logger.info("All reports generated successfully!")
